import datetime
import math
import threading
import time
import turtle
import colorsys
import numpy as np
import functools
import cv2
import mediapipe as mp
import pyautogui
import xlsxwriter
import openpyxl
from pylsl import StreamInlet, resolve_stream
import statistics
# Global flag to control the webcam interaction loop
stop_webcam = False

data = []  # List to store data points
modeData = []  # List to store modes (not used in the provided code)
count = 0  # Counter variable
timeData = 0
# Initialize camera
cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
dot_positions = [int] * 9

# Class for creating blinking lights using Turtle graphics
class BlinkingLights:
    def __init__(self, data_list , statsData):
        # Initialize Turtle graphics window
        self.screen = turtle.Screen()
        self.screen.title('Blinking Lights')
        self.screenTk = self.screen.getcanvas().winfo_toplevel()
        self.screenTk.attributes("-fullscreen", True)
        self.screen.bgcolor("black")
        self.round = 0  # Counter for rounds of blinking
        self.brightness = 0.0  # Initial brightness of lights
        self.hues = [0.0, 0.2, 0.4, 0.6, 0.8, 1.0, 1.2, 1.4, 1.6]  # List of hues for colors
        self.dots = []  # List to store Turtle objects representing lights
        self.blinking = []  # List to store blinking status of each light
        self.data_list = data_list  # Reference to the shared data list
        self.statsData=statsData
        self.workbook = xlsxwriter.Workbook("CalibrationData.xlsx")  # Workbook for storing data
        self.worksheet = self.workbook.add_worksheet("firstSheet")  # Worksheet for storing data
        # Write column headers in the worksheet
        self.worksheet.write(0, 0, "Top Left (x,y,time stamp)")
        self.worksheet.write(0, 1, "Top Center (x,y,time stamp)")
        self.worksheet.write(0, 2, "Top Right (x,y,time stamp)")
        self.worksheet.write(0, 3, "Center Left (x,y,time stamp)")
        self.worksheet.write(0, 4, "Center (x,y,time stamp)")
        self.worksheet.write(0, 5, "Center Right (x,y,time stamp)")
        self.worksheet.write(0, 6, "Bottom Left (x,y,time stamp)")
        self.worksheet.write(0, 7, "Bottom Center (x,y,time stamp)")
        self.worksheet.write(0, 8, "Bottom Right (x,y,time stamp)")
        self.row = 0  # Row index for writing data
        self.col = 0  # Column index for writing data

    @staticmethod
    def create_dot(position):
        # Create a Turtle object representing a light dot
        dot = turtle.Turtle()
        dot.shape("circle")
        dot.shapesize(1, 1)
        dot.goto(position)
        dot.penup()
        return dot

    def change_brightness_sequence(self, index):
        # Method to change the brightness of lights in sequence
        self.round += 1
        global count
        count = self.round
        if self.round > 9:
            # If completed 9 rounds, close the Turtle graphics window and workbook
            time.sleep(3)
            self.screen.bye()
            self.workbook.close()

        # Toggle brightness between 0.1 and 0.9
        self.brightness = 0.9 if self.brightness < 1.0 else 0.1
        rgb = colorsys.hsv_to_rgb(self.hues[index], 1, self.brightness)
        self.dots[index].color(rgb)

        x_mode = statistics.mode( [x[0] for x in statsData])
        y_mode = statistics.mode([y[1] for y in statsData])
        modeData.append([x_mode ,y_mode])
        dot_position = list(self.dots[index].position())
        #dot_positions.append(dot_position)
        statsData.clear()
        # Write data to the worksheet for each light
        column = 1
        for line in self.data_list:
            self.worksheet.write(column, self.round - 1, line)
            column += 1
        # Clear the shared data list
        data_list.clear()

        # Schedule the next change in brightness
        if self.blinking[index]:
            self.screen.ontimer(functools.partial(self.change_brightness_sequence, index), 100)
            self.data_list.append(self.dots[index].position())
        else:
            self.screen.ontimer(functools.partial(self.hide_dot, index), 1000)

    def hide_dot(self, index):
        # Method to hide a light dot and move to the next dot
        self.dots[index].hideturtle()
        next_index = (index + 1) % len(self.dots)
        self.dots[next_index] = self.create_dot(self.dots[next_index].position())
        self.screen.ontimer(functools.partial(self.change_brightness_sequence, next_index), 200)

    def start_blinking_lights(self):
        instruction = turtle.Turtle()
        instruction.hideturtle()
        instruction.penup()
        instruction.color("white")  # Set text color to white
        instruction.goto(0, 0)
        instruction.write("Please follow the color dot on the screen for calibration.", align="center", font=("Arial", 24, "normal"))
        self.screen.update()  # Update the screen to display the instruction

        # Schedule the removal of instruction text after 5 seconds
        self.screen.ontimer(lambda: instruction.clear(), 5000)

        # Create light dots at specific positions
        self.dots = [self.create_dot((-800 + (i % 3) * 800, 500 - int(i / 3) * 500)) for i in range(9)]
        self.blinking = [False] * len(self.dots)

        # Start the sequence by changing brightness of the first light after 5 seconds
        self.screen.ontimer(functools.partial(self.change_brightness_sequence, 0), 5000)
        self.screen.mainloop()


# Function to start the Turtle graphics for blinking lights
def start_turtle_graphics(data_list , statsData):
    blinking_lights = BlinkingLights(data_list,statsData)
    blinking_lights.start_blinking_lights()

def distance(point1, point2):
    return math.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

def closest_coordinate(point, coordinates):
    min_distance = float('inf')
    closest_coord = None
    for coord in coordinates:
        dist = distance(point, coord)
        if dist < min_distance:
            min_distance = dist
            closest_coord = coord
    return closest_coord

# Function to start webcam interaction for detecting facial landmarks
def start_webcam_interaction(data_list=None , statsData=None):
    global stop_webcam  # Access the global flag
    face_mesh = mp.solutions.face_mesh.FaceMesh(refine_landmarks=True)
    screen_w, screen_h = pyautogui.size()
    dim = (screen_w, screen_h)
    f_sc = cv2.VideoWriter_fourcc(*'XVID')
    out_sc = cv2.VideoWriter('screen_recording.mp4',f_sc,60.0,dim)
    workbook = xlsxwriter.Workbook("AfterCalibration.xlsx")
    worksheet = workbook.add_worksheet("firstSheet")
    worksheet.write(0, 0, "Point 1")
    row = 0
    col = 0
    dot_positions[0]=[(screen_h/3)-40,(screen_w/3)-80]
    dot_positions[1]=[(screen_h/3)-40,(screen_w/3)*2-80]
    dot_positions[2]=[(screen_h/3)-40,(screen_w/3)*3-80]
    dot_positions[3]=[(screen_h/3)*2-40,(screen_w/3)-80]
    dot_positions[4]=[(screen_h/3)*2-40,(screen_w/3)*2-80]
    dot_positions[5]=[(screen_h/3)*2-40,(screen_w/3)*3-80]
    dot_positions[6]=[(screen_h/3)-40,(screen_w/3)-80]
    dot_positions[7]=[(screen_h/3)-40,(screen_w/3)*2-80]
    dot_positions[8]=[(screen_h/3)-40,(screen_w/3)*3-80]

    print(dot_positions)
    while not stop_webcam:  # Continue the loop until the stop flag is set
        ret, frame = cam.read()
        if not ret:
            print("Error: Unable to access the camera.")
            break
        frame = cv2.flip(frame, 1)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        output = face_mesh.process(rgb_frame)
        landmark_points = output.multi_face_landmarks
        frame_h, frame_w, ret = frame.shape
        if landmark_points:
            landmarks = landmark_points[0].landmark
            for id, landmark in enumerate(landmarks[474:478]):
                x = int(landmark.x * frame_w)
                y = int(landmark.y * frame_h)
                cv2.circle(frame, (x, y), 3, (0, 255, 0))
                if id == 1:
                    screen_x = int(screen_w * landmark.x)
                    screen_y = int(screen_h * landmark.y)
                    current_time = datetime.datetime.now()
                    global timeData
                    timeData = current_time
                    if data_list is not None:
                        data_list.append('X: ' + str(screen_x) + ' Y: ' + str(screen_y) + " Time:  " + str(current_time))
                        statsData.append([screen_x,screen_y])
                    if count > 9:
                        im_sc = pyautogui.screenshot()
                        fr_sc = np.array(im_sc)
                        point = [screen_x,screen_y]

                        closest_coord = closest_coordinate(point,modeData)
                        closestCordIndex = modeData.index(closest_coord)
                        dotPositionData = dot_positions[closestCordIndex]

                        cv2.circle(fr_sc, (int(dotPositionData[0]), int(dotPositionData[1])), 50, (0, 255, 255))
                        rgb_sc = cv2.cvtColor(fr_sc, cv2.COLOR_BGR2RGB)
                        out_sc.write(rgb_sc)
                        worksheet.write(col, row,
                                        'X: ' + str(screen_x) + ' Y: ' + str(screen_y) + " Time: " + str(
                                            current_time))
                        col += 1
            left = [landmarks[145], landmarks[159]]
            for landmark in left:
                x = int(landmark.x * frame_w)
                y = int(landmark.y * frame_h)
                cv2.circle(frame, (x, y), 3, (0, 255, 0))
            if (left[0].y - left[1].y) < 0.004:
                pyautogui.click()
                pyautogui.sleep(1)

        cv2.imshow('Eye Controlled ', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            workbook.close()
            break

    # Release the camera when done
    out_sc.release()
    cam.release()
    cv2.destroyAllWindows()


def lsl_streaming():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ["Timestamp", "Data"]
    sheet.append(header)
    workbook.save("LSL_Data.xlsx")
    # first resolve an EEG stream on the lab network
    print("looking for an Event stream...")
    streams = resolve_stream('type', 'Event')

    # create a new inlet to read from the stream
    inlet = StreamInlet(streams[0])

    while True:
        # get a new sample (you can also omit the timestamp part if you're not
        # interested in it)
        sample, timestamp = inlet.pull_sample()
        #print(timestamp, sample)

        # Write data to the worksheet
        for col, data_point in enumerate(sample):
            sheet.append([str(data_point) + " TimeStamp: " + str(timeData)])
        workbook.save("LSL_Data.xlsx")
        # Sleep for a short time to avoid excessive CPU usage
        time.sleep(0.1)


if __name__ == "__main__":
    data_list = []  # Shared list for storing data points
    statsData = []
    # Create threads for running Turtle graphics, webcam interaction, and LSL streaming concurrently
    turtle_thread = threading.Thread(target=start_turtle_graphics, args=(data_list, statsData))
    webcam_thread = threading.Thread(target=start_webcam_interaction, args=(data_list , statsData))
    lsl_thread = threading.Thread(target=lsl_streaming, args=())
    # Start the threads
    turtle_thread.start()
    webcam_thread.start()
    lsl_thread.start()
    # Wait for threads to finish
    turtle_thread.join()
    # Set the flag to stop the webcam interaction loop
    #stop_webcam = True
    webcam_thread.join()
    lsl_thread.join()