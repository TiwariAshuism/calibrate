import datetime
import math
import statistics
import threading
import time
import cv2
import mediapipe as mp
import numpy as np
import openpyxl
import pyautogui
import xlsxwriter
from openpyxl.workbook import Workbook
from pylsl import StreamInlet, resolve_stream

cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
data = []
modeData = []
count = 0
timeData = 0
blinking_index = 0
dot_location = []


def draw_dot(frame, center, radius, color):
    cv2.circle(frame, center, radius, color, thickness=-1)


def instruction():
    screen_resolution = pyautogui.size()
    fra = np.zeros((screen_resolution[1], screen_resolution[0], 3), dtype=np.uint8)
    fra = cv2.putText(fra, 'Please follow the nine green dots on the screen.',
                      ((screen_resolution[0] // 2)-200 , screen_resolution[1] // 2), cv2.FONT_HERSHEY_SIMPLEX, 1,
                      (0, 255, 0), 2)

    cv2.namedWindow("Instructions", cv2.WINDOW_NORMAL)
    cv2.setWindowProperty("Instructions", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
    cv2.imshow("Instructions", fra)
    start_time = time.time()
    while time.time() - start_time < 5:
        if cv2.waitKey(10) & 0xFF == ord('q'):
            break
    cv2.destroyWindow("Instructions")


start_time = time.time()


def main(data_list):
    current_time = datetime.datetime.now()
    screen_resolution = (1920, 1080)
    dot_radius = 10
    dot_color = (0, 255, 0)

    dot_positions = [(screen_resolution[0] // 6, screen_resolution[1] // 6),
                     (screen_resolution[0] // 2, screen_resolution[1] // 6),
                     (5 * screen_resolution[0] // 6, screen_resolution[1] // 6),
                     (screen_resolution[0] // 6, screen_resolution[1] // 2),
                     (screen_resolution[0] // 2, screen_resolution[1] // 2),
                     (5 * screen_resolution[0] // 6, screen_resolution[1] // 2),
                     (screen_resolution[0] // 6, 5 * screen_resolution[1] // 6),
                     (screen_resolution[0] // 2, 5 * screen_resolution[1] // 6),
                     (5 * screen_resolution[0] // 6, 5 * screen_resolution[1] // 6)]

    dot_positions_transposed = list(map(list, zip(*dot_positions)))
    global dot_location
    for i in range(len(dot_positions_transposed[0])):
        dot_location.append([dot_positions_transposed[0][i], dot_positions_transposed[1][i]])
    wb = Workbook()
    ws = wb.active
    global modeData
    cv2.namedWindow("Blinking Dots", cv2.WINDOW_NORMAL)
    cv2.setWindowProperty("Blinking Dots", cv2.WND_PROP_TOPMOST, 1)

    global blinking_index

    while blinking_index < 9:
        frame = np.zeros((screen_resolution[1], screen_resolution[0], 3), dtype=np.uint8)
        draw_dot(frame, dot_positions[blinking_index], dot_radius, dot_color)
        cv2.imshow("Blinking Dots", frame)
        blinking_index += 1
        start_time = time.time()
        initial_length = len(data_list)  # Capture initial length

        while time.time() - start_time < 1:
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        # Clear the data_list after processing
        x_mode = statistics.mode([x[0] for x in data_list])
        y_mode = statistics.mode([y[1] for y in data_list])
        ws.cell(row=1, column=blinking_index).value = f"X Mode: {x_mode}, Y Mode: {y_mode}"
        for i, row in enumerate(data_list):
            ws.cell(row=i + 2,
                    column=blinking_index).value = f" X: {row[0]}, Y: {row[1]} , Timestamp: {current_time}"  # Write X data
        wb.save('dot_positions.xlsx')
        modeData.append([x_mode, y_mode])
        print(data_list)
        del data_list[initial_length:]

    cv2.destroyWindow("Blinking Dots")


def distance(point1, point2):
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)


def closest_coordinate(point, coordinates):
    min_distance = float('inf')
    closest_coord = None
    for coord in coordinates:
        dist = distance(point, coord)
        if dist < min_distance:
            min_distance = dist
            closest_coord = coord
    return closest_coord


def start_webcam_interaction(data_list=None):
    global dot_location
    global blinking_index

    face_mesh = mp.solutions.face_mesh.FaceMesh(refine_landmarks=True)
    screen_w, screen_h = pyautogui.size()
    dim = (screen_w, screen_h)
    f_sc = cv2.VideoWriter_fourcc(*'XVID')
    out_sc = cv2.VideoWriter('screen_recording.mp4', f_sc, 10.0, dim)
    workbook = xlsxwriter.Workbook("AfterCalibration.xlsx")
    worksheet = workbook.add_worksheet("firstSheet")
    worksheet.write(0, 0, "Point 1")
    row = 0
    col = 0
    while True:
        ret, frame = cam.read()
        if not ret:
            print("Error: Unable to access the camera.")
            break
        frame = cv2.flip(frame, 1)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        output = face_mesh.process(rgb_frame)
        landmark_points = output.multi_face_landmarks
        frame_h, frame_w, ret = frame.shape
        if landmark_points:
            landmarks = landmark_points[0].landmark
            for id, landmark in enumerate(landmarks[474:478]):
                x = int(landmark.x * frame_w)
                y = int(landmark.y * frame_h)
                cv2.circle(frame, (x, y), 3, (0, 255, 0))
                if id == 1:
                    screen_x = int(screen_w * landmark.x)
                    screen_y = int(screen_h * landmark.y)
                    current_time = datetime.datetime.now()
                    global timeData
                    timeData = current_time
                    point = [screen_x, screen_y]

                    if data_list is not None and blinking_index < 9:
                        data_list.append(point)
                    if blinking_index >= 9:
                        im_sc = pyautogui.screenshot()
                        fr_sc = np.array(im_sc)
                        closest_coord = closest_coordinate(point, modeData)
                        closestCordIndex = modeData.index(closest_coord)
                        dotPositionData = dot_location[closestCordIndex]
                        cv2.circle(fr_sc, (dotPositionData[0], dotPositionData[1]), 150, (255, 0, 0))
                        rgb_sc = cv2.cvtColor(fr_sc, cv2.COLOR_BGR2RGB)
                        out_sc.write(rgb_sc)
                        worksheet.write(col, row,
                                        'X: ' + str(screen_x) + ' Y: ' + str(screen_y) + " Time: " + str(
                                            current_time))
                        col += 1
            left = [landmarks[145], landmarks[159]]
            for landmark in left:
                x = int(landmark.x * frame_w)
                y = int(landmark.y * frame_h)
                cv2.circle(frame, (x, y), 3, (0, 255, 0))
            if (left[0].y - left[1].y) < 0.004:
                pyautogui.click()
                pyautogui.sleep(1)
        cv2.imshow('Eye Controlled ', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    workbook.close()
    out_sc.release()
    cam.release()
    cv2.destroyAllWindows()


def lsl_streaming():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ["Timestamp", "Data"]
    sheet.append(header)
    workbook.save("LSL_Data.xlsx")
    print("looking for an Event stream...")
    streams = resolve_stream('type', 'Event')
    inlet = StreamInlet(streams[0])

    while True:
        sample, timestamp = inlet.pull_sample()
        for col, data_point in enumerate(sample):
            sheet.append([str(data_point) + " TimeStamp: " + str(timeData)])
        workbook.save("LSL_Data.xlsx")
        time.sleep(0.1)


if __name__ == "__main__":
    data_list = []  # Shared list for storing data points
    statsData = []
    instructionThread = threading.Thread(target=instruction, args=())
    instructionThread.start()
    instructionThread.join()
    time.sleep(5)
    # Create threads for running main, webcam interaction, and LSL streaming concurrently
    main_thread = threading.Thread(target=main, args=(data_list,))
    webcam_thread = threading.Thread(target=start_webcam_interaction, args=(data_list,))
    lsl_thread = threading.Thread(target=lsl_streaming, args=())

    # Start the threads
    main_thread.start()
    webcam_thread.start()
    lsl_thread.start()

    # Wait for threads to finish
    main_thread.join()
    webcam_thread.join()
    lsl_thread.join()
